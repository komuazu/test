# -*- coding: utf-8 -*-
"""
稼動日報 → 営業部別 印刷実績 自動集計

使い方:
    python kadou_report.py --month 2025-09 --out ./出力 日報1.xls 日報2.xls ...
    python kadou_report.py --month 9 --out ./出力 *.xls        # 年を省略すると最新年を推定

処理内容（2026-08 時点の確定ルール）:
  1) 各 .xls から「25年9月」形式の月次シートを探し、明細行と日次集計行を抽出
     （「25年9月30」のような日付付きの補助シートは対象外）
  2) 営業担当ｺｰﾄﾞで部門に振り分け
        本社営業部 = 1110, 1120 ／ 東京営業部 = 2100, 2140 ／ 池袋営業部 = 3810, 3820
        （上記以外は出力対象外）
  3) 管理番号ごとに統合。枝番（-1 / -1-2 / -2 …）は同一管理番号として統合
        通し数 = 合計 ／ 色数 = 「表/裏」(表版数・裏版数の最大値)
        印刷日 = 初日 ／ 品名 = 初出の1つ ／ ｸﾗｲｱﾝﾄ名 = 最初に見つかった非空値
  4) 並び替え: ①営業担当ｺｰﾄﾞ降順 ②ｸﾗｲｱﾝﾄ名降順(NFKC同一視・空欄は末尾) ③印刷日降順
  5) 部門ごとに1ファイルずつ出力（罫線・入力欄3列・合計/差引行つき）
"""
import argparse, datetime, re, sys, unicodedata
from collections import OrderedDict
from pathlib import Path

try:
    import xlrd
except ImportError:
    sys.exit('xlrd が必要です:  pip install xlrd')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ────────── 設定 ──────────
DEPTS = OrderedDict([('本社営業部', [1110, 1120]),
                     ('東京営業部', [2100, 2140]),
                     ('池袋営業部', [3810, 3820])])
CODE2DEPT = {c: d for d, cs in DEPTS.items() for c in cs}

HEAD = ['管理番号', 'ｸﾗｲｱﾝﾄ名', '品名', '今年の動向', '無しの場合の代替対策',
        '対策通し数', '営業担当ｺｰﾄﾞ', '印刷日', '色数', '通し数']
INPUT_COLS = (4, 5, 6)
WIDTHS = [14, 24, 42, 26, 32, 13, 14, 12, 10, 12]

# 必要な列は見出し文字で動的に特定する（月によって列位置が違うため）
NEED = ('日付', '管理番号', '営業担当ｺｰﾄﾞ', '品名', '通し枚数', '表版数', '裏版数')
OPTIONAL = ('ｸﾗｲｱﾝﾄ名',)
TIME_LABELS = ('有効時間', '準備合計', '色合わせ', '印刷時間', 'その他')
EPOCH = datetime.date(1899, 12, 30)

# 日本語グリフを持つフォントにする（Arial など欧文専用フォントを日本語セルに指定すると、
# 印刷・印刷プレビューでのフォント差し替えが Excel の異常終了を招くことがある）
FONT = 'Meiryo'
C_HDR, C_INHDR = '1F3864', 'BF8F00'
C_IN, C_TOT, C_DIF, C_MD, C_LEG = 'FFFF00', 'FFF2CC', 'FCE4D6', 'FFF9DB', 'FFF9DB'
TH = Side(style='thin', color='808080')
MD = Side(style='medium', color='1F3864')
HW = Side(style='thin', color='FFFFFF')


# ────────── ユーティリティ ──────────
def norm(s):
    return str(s).replace('　', '').strip()


def ckey(s):
    if not s:
        return ''
    return unicodedata.normalize('NFKC', str(s)).replace(' ', '').replace('　', '')


class Desc:
    """昇順ソートの中で文字列だけ降順にするラッパ"""
    __slots__ = ('v',)
    def __init__(self, v): self.v = v
    def __lt__(self, o): return self.v > o.v
    def __eq__(self, o): return self.v == o.v


def basno(no):
    """管理番号から枝番を落とした基番号を返す  8632175-1-2 → 8632175"""
    s = str(no).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s.split('-')[0]


SHEET_RE = re.compile(r'^(?:20)?(\d{2})\s*[年月]\s*(\d{1,2})\s*月$')


def find_month_sheet(wb, year2, month):
    """「25年9月」形式の月次シート名を返す。日付付き補助シートは除外。"""
    for name in wb.sheet_names():
        m = SHEET_RE.match(norm(name))
        if m and int(m.group(1)) == year2 and int(m.group(2)) == month:
            return name
    return None


def cell(sh, r, c):
    if c is None or c >= sh.ncols:
        return None
    v = sh.cell_value(r, c)
    return None if v == '' else v


def header_map(sh):
    """見出し行を探し、{列名: 最初に現れた列番号} を返す。
    「通し枚数」「表版数」等は本刷ブロックと集計ブロックの2か所に出るため最初の方を採用。"""
    for hr in range(0, min(8, sh.nrows)):
        names = [norm(sh.cell_value(hr, c)) for c in range(sh.ncols)]
        if '日付' in names and '管理番号' in names:
            m = {}
            for c, n in enumerate(names):
                if n and n not in m:
                    m[n] = c
            missing = [k for k in NEED if k not in m]
            if missing:
                raise ValueError('見出し列が見つかりません: ' + ', '.join(missing))
            for k in OPTIONAL:
                m.setdefault(k, None)
            return hr, m
    raise ValueError('見出し行（「日付」「管理番号」を含む行）が見つかりません')


# ────────── 抽出 ──────────
def extract(path, year2, month):
    """1ファイルから明細行のリストを返す"""
    wb = xlrd.open_workbook(str(path))
    name = find_month_sheet(wb, year2, month)
    if not name:
        return None, []
    sh = wb.sheet_by_name(name)
    hr, M = header_map(sh)
    out = []
    for r in range(hr + 1, sh.nrows):
        if norm(sh.cell_value(r, 27) if sh.ncols > 27 else '') in TIME_LABELS + ('受注件数',):
            continue                                    # 日次集計ブロックは明細ではない
        d = cell(sh, r, M['日付'])
        if not (isinstance(d, float) and d > 40000):     # 日付シリアルの行だけが明細
            continue
        code = cell(sh, r, M['営業担当ｺｰﾄﾞ'])
        code = int(code) if isinstance(code, float) else code
        if code not in CODE2DEPT:
            continue
        out.append({
            'dept':   CODE2DEPT[code],
            'code':   code,
            'no':     cell(sh, r, M['管理番号']),
            'base':   basno(cell(sh, r, M['管理番号'])),
            'date':   EPOCH + datetime.timedelta(days=int(d)),
            'client': cell(sh, r, M['ｸﾗｲｱﾝﾄ名']),
            'name':   cell(sh, r, M['品名']),
            'f':      cell(sh, r, M['表版数']) or 0,
            'b':      cell(sh, r, M['裏版数']) or 0,
            'tsu':    cell(sh, r, M['通し枚数']) or 0,
            'seq':    r,
        })
    return name, out


def consolidate(rows):
    """管理番号（枝番統合）ごとにまとめる。
    ｸﾗｲｱﾝﾄ名・品名の代表値は入力ファイルの順序に依存しない決定的ルールで選ぶ。"""
    g = OrderedDict()
    for r in rows:
        k = (r['dept'], r['base'])
        e = g.setdefault(k, {'dept': r['dept'], 'base': r['base'], 'code': r['code'],
                             'dates': [], 'f': 0, 'b': 0, 'tsu': 0, 'n': 0,
                             'nos': set(), 'members': []})
        e['members'].append(r)
        e['dates'].append(r['date'])
        e['nos'].add(str(r['no']))
        e['f'] = max(e['f'], r['f'])
        e['b'] = max(e['b'], r['b'])
        e['tsu'] += r['tsu']
        e['n'] += 1
    for e in g.values():
        # ｸﾗｲｱﾝﾄ名: 非空値の最頻値 → 同数なら最も長い（情報量が多い）→ なお同じなら文字列順
        names = [str(m['client']) for m in e['members'] if m['client']]
        e['client'] = (max(set(names), key=lambda s: (names.count(s), len(s), s))
                       if names else None)
        # 品名: 初日の行 → 管理番号 → 品名 の順で最小のもの
        first = min(e['members'], key=lambda m: (m['date'], str(m['no']), str(m['name'] or '')))
        e['name'] = first['name']
        del e['members']
    return list(g.values())


def sort_groups(gs):
    gs.sort(key=lambda g: (-g['code'],
                           g['client'] in (None, ''),
                           Desc(ckey(g['client'])),
                           -min(g['dates']).toordinal(),
                           g['base']))
    return gs


# ────────── 出力 ──────────
def write_sheet(ws, dept, gs, year, month, src_rows, src_files):
    codes = ', '.join(map(str, DEPTS[dept]))
    multi_d = sum(1 for g in gs if len(set(g['dates'])) > 1)
    multi_n = sum(1 for g in gs if len(g['nos']) > 1)
    ws['A1'] = '%s ／ %d年%d月 印刷実績（管理番号ごとに統合・枝番も統合）' % (dept, year, month)
    ws['A1'].font = Font(name=FONT, size=13, bold=True)
    ws['A2'] = ('対象 営業担当ｺｰﾄﾞ: %s ／ 明細%d行 → %d件に統合（枝番統合 %d件）。'
                '通し数=合計、色数=「表/裏」（表版数・裏版数の最大値）、印刷日=初日'
                '（複数日にまたがる%d件は黄色）、品名=初出の1つ。'
                '「今年の動向」「無しの場合の代替対策」「対策通し数」は黄色セルの手入力欄。'
                '並び順: ①営業担当ｺｰﾄﾞ降順 ②ｸﾗｲｱﾝﾄ名降順（NFKCで表記ゆれを同一視／空欄は末尾） ③印刷日降順。'
                '元ファイル: %s') % (codes, src_rows, len(gs), multi_n, multi_d, ' / '.join(src_files))
    ws['A2'].font = Font(name=FONT, size=9, italic=True, color='595959')

    for c, h in enumerate(HEAD, 1):
        x = ws.cell(row=4, column=c, value=h)
        x.font = Font(name=FONT, size=10, bold=True, color='FFFFFF')
        x.fill = PatternFill('solid', fgColor=C_INHDR if c in INPUT_COLS else C_HDR)
        x.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[4].height = 30

    r0 = 5
    for i, g in enumerate(gs):
        r = r0 + i
        vals = [g['base'], g['client'], g['name'], None, None, None,
                g['code'], min(g['dates']), '%d/%d' % (g['f'], g['b']), g['tsu']]
        for c, v in enumerate(vals, 1):
            x = ws.cell(row=r, column=c, value=v)
            if c in INPUT_COLS:
                x.font = Font(name=FONT, size=10, color='0000FF')
                x.fill = PatternFill('solid', fgColor=C_IN)
                x.alignment = (Alignment(horizontal='right') if c == 6
                               else Alignment(vertical='top', wrap_text=True))
                if c == 6:
                    x.number_format = '#,##0'
            else:
                x.font = Font(name=FONT, size=10)
                if c == 7:
                    x.number_format = '0'; x.alignment = Alignment(horizontal='center')
                elif c == 8:
                    x.number_format = 'yyyy/mm/dd'
                    if len(set(g['dates'])) > 1:
                        x.fill = PatternFill('solid', fgColor=C_MD)
                elif c == 9:
                    x.alignment = Alignment(horizontal='center')
                elif c == 10:
                    x.number_format = '#,##0'
        if len(g['nos']) > 1:
            ws.cell(row=r, column=1).font = Font(name=FONT, size=10, bold=True)

    r1, tr = r0 + len(gs) - 1, r0 + len(gs)
    er = tr + 1
    ws.cell(row=tr, column=1, value='合計')
    ws.cell(row=tr, column=3, value='%d件' % len(gs))
    ws.cell(row=tr, column=6, value='=SUM(F%d:F%d)' % (r0, r1)).number_format = '#,##0'
    ws.cell(row=tr, column=10, value='=SUM(J%d:J%d)' % (r0, r1)).number_format = '#,##0'
    ws.cell(row=er, column=1, value='差引（通し数 − 対策通し数）')
    ws.cell(row=er, column=6, value='=J%d-F%d' % (tr, tr)).number_format = '#,##0'
    ws.cell(row=er, column=9, value='充足率→').alignment = Alignment(horizontal='right')
    ws.cell(row=er, column=10, value='=IF(J%d=0,"",F%d/J%d)' % (tr, tr, tr)).number_format = '0.0%'
    for r, fill in ((tr, C_TOT), (er, C_DIF)):
        for c in range(1, 11):
            x = ws.cell(row=r, column=c)
            x.font = Font(name=FONT, size=10, bold=True)
            x.fill = PatternFill('solid', fgColor=fill)

    # 罫線
    for c in range(1, 11):
        ws.cell(row=4, column=c).border = Border(left=MD if c == 1 else HW,
                                                 right=MD if c == 10 else HW, top=MD, bottom=MD)
    for r in range(r0, r1 + 1):
        for c in range(1, 11):
            ws.cell(row=r, column=c).border = Border(left=MD if c == 1 else TH,
                                                     right=MD if c == 10 else TH,
                                                     top=TH, bottom=MD if r == r1 else TH)
    for r in (tr, er):
        for c in range(1, 11):
            ws.cell(row=r, column=c).border = Border(left=MD if c == 1 else TH,
                                                     right=MD if c == 10 else TH,
                                                     top=MD if r == tr else TH, bottom=MD)
    # 凡例と記入例
    lr = er + 2
    ws.cell(row=lr, column=1, value='【記入方法】').font = Font(name=FONT, size=10, bold=True)
    for i, (k, t) in enumerate([
            ('今年の動向', '今年その商材がどうなるか（例: 継続／部数減／今年は無し／電子化 など）'),
            ('無しの場合の代替対策', '無くなる・減る場合に何で埋めるか（例: 他商材の前倒し／新規A社案件 など）'),
            ('対策通し数', '代替対策で見込める通し数（数値）。合計と差引・充足率が上の行に自動計算されます')]):
        ws.cell(row=lr + 1 + i, column=1, value=k).font = Font(name=FONT, size=9, bold=True)
        ws.cell(row=lr + 1 + i, column=3, value=t).font = Font(name=FONT, size=9, color='595959')
        for c in range(1, 4):
            ws.cell(row=lr + 1 + i, column=c).border = Border(left=TH, right=TH, top=TH, bottom=TH)
    for c in range(1, 4):
        ws.cell(row=lr, column=c).border = Border(left=TH, right=TH, top=TH, bottom=TH)
    ex = lr + 5
    ws.cell(row=ex, column=1, value='記入例 →').font = Font(name=FONT, size=9, bold=True, color=C_INHDR)
    for c, v in ((4, '部数減（前年比 約70%）'), (5, 'B社カタログの前倒し受注で充当'), (6, 12000)):
        x = ws.cell(row=ex, column=c, value=v)
        x.font = Font(name=FONT, size=9, italic=True, color=C_INHDR)
        x.fill = PatternFill('solid', fgColor=C_LEG)
        if c == 6:
            x.number_format = '#,##0'
    for c in range(1, 7):
        ws.cell(row=ex, column=c).border = Border(left=TH, right=TH, top=TH, bottom=TH)
    ws.cell(row=ex + 1, column=1,
            value='※ 記入例はサンプルです。入力は上の黄色いセルへ。管理番号が太字の行は複数の枝番を統合しています（内訳は下の一覧）。'
            ).font = Font(name=FONT, size=9, italic=True, color='595959')

    # 枝番の内訳
    # （以前はセルのコメントに入れていたが、コメントは図形レイヤー＝印刷時に描画される部分に
    #   置かれるため、印刷・印刷プレビューでの不具合の原因になりうる。表として書き出す）
    merged = [g for g in gs if len(g['nos']) > 1]
    if merged:
        mr = ex + 3
        ws.cell(row=mr, column=1, value='【統合した管理番号の内訳】').font = Font(name=FONT, size=10, bold=True)
        ws.cell(row=mr, column=3,
                value='管理番号が太字の行は、次の枝番をまとめた合計です。'
                ).font = Font(name=FONT, size=9, color='595959')
        for i, g in enumerate(merged):
            ws.cell(row=mr + 1 + i, column=1, value=g['base']).font = Font(name=FONT, size=9, bold=True)
            ws.cell(row=mr + 1 + i, column=3, value=', '.join(sorted(g['nos']))
                    ).font = Font(name=FONT, size=9)
        for rr in range(mr, mr + len(merged) + 1):
            for c in range(1, 4):
                ws.cell(row=rr, column=c).border = Border(left=TH, right=TH, top=TH, bottom=TH)

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = 'A4:J%d' % r1
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.zoomScale = 100
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4      # 用紙をA4に固定（プリンタ既定に左右されない）
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    # 見出し行(4行目)を各ページの先頭で繰り返す。
    # 繰り返す行が印刷範囲の中にあると1ページ目で二重に印刷されるため、印刷範囲は5行目から。
    ws.print_title_rows = '4:4'
    ws.print_area = 'A5:J%d' % er
    # 1〜3行目は印刷範囲外になるので、表題と条件はページヘッダー／フッターで各ページに出す
    ws.oddHeader.left.text = '%s ／ %d年%d月 印刷実績' % (dept, year, month)
    ws.oddHeader.left.font, ws.oddHeader.left.size = FONT, 10
    ws.oddHeader.right.text = '&P / &N'
    ws.oddHeader.right.font, ws.oddHeader.right.size = FONT, 9
    ws.oddFooter.left.text = ('管理番号ごとに統合・枝番も統合 ／ 通し数=合計 ／ '
                              '色数=表/裏 ／ 印刷日=初日')
    ws.oddFooter.left.font, ws.oddFooter.left.size = FONT, 8


def run(files, year, month, outdir):
    year2 = year % 100
    rows, used, missing = [], [], []
    for f in files:
        name, got = extract(f, year2, month)
        if name is None:
            missing.append(Path(f).name)
        else:
            used.append(Path(f).name)
            rows += got
    if missing:
        print('  [注意] 対象月のシートが見つからないファイル: ' + ', '.join(missing))
    if not rows:
        raise SystemExit('対象月（%d年%d月）の該当データが1件も見つかりませんでした。' % (year, month))
    rows.sort(key=lambda x: (x['date'], str(x['no']), x['seq']))

    outdir = Path(outdir); outdir.mkdir(parents=True, exist_ok=True)
    results = []
    for dept in DEPTS:
        drows = [r for r in rows if r['dept'] == dept]
        gs = sort_groups(consolidate(drows))
        wb = Workbook(); ws = wb.active; ws.title = dept
        write_sheet(ws, dept, gs, year, month, len(drows), used)
        path = outdir / ('%d年%d月_%s_印刷実績.xlsx' % (year, month, dept))
        wb.save(path)
        results.append((dept, len(drows), len(gs), sum(g['tsu'] for g in gs), path))
    return results


def main():
    p = argparse.ArgumentParser(description='稼動日報から営業部別の印刷実績を作成します')
    p.add_argument('files', nargs='+', help='稼動日報の .xls ファイル')
    p.add_argument('--month', required=True, help='対象月  例: 2025-09 / 2025/9 / 9')
    p.add_argument('--out', default='.', help='出力先フォルダ')
    a = p.parse_args()
    m = re.match(r'^(?:(\d{4})[-/年])?\s*(\d{1,2})月?$', a.month.strip())
    if not m:
        raise SystemExit('--month の書式が不正です（例: 2025-09 / 9）')
    year = int(m.group(1)) if m.group(1) else datetime.date.today().year
    month = int(m.group(2))
    print('対象: %d年%d月 ／ 入力%d件' % (year, month, len(a.files)))
    for dept, n, g, tsu, path in run(a.files, year, month, a.out):
        print('  %-6s 明細%4d行 → %3d件  通し数 %10s  →  %s' % (dept, n, g, format(int(tsu), ','), path.name))
    print('完了')


if __name__ == '__main__':
    main()
